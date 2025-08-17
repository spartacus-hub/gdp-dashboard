#!/usr/bin/env python3

from __future__ import annotations

import os
from io import BytesIO
from typing import Dict, List, Tuple

import streamlit as st

# Soft import: show a friendly message if missing at runtime
try:
	import pandas as pd
except Exception as e:
	pd = None

try:
	import plotly.express as px
except Exception as e:
	px = None

st.set_page_config(page_title="Tableau de bord Excel", layout="wide")

# --------------- Helpers ---------------

def ensure_deps() -> None:
	missing = []
	if pd is None:
		missing.append("pandas")
	if px is None:
		missing.append("plotly")
	if missing:
		st.error(
			"Les dépendances suivantes sont manquantes: " + ", ".join(missing) + 
			". Installez-les puis relancez l'application."
		)
		st.stop()


@st.cache_data(show_spinner=False)
def read_excel_to_dataframe(file_bytes: bytes, sheet_name: str | None = None) -> pd.DataFrame:
	"""Read Excel bytes to a pandas DataFrame using openpyxl engine."""
	excel_io = BytesIO(file_bytes)
	# Let pandas infer the sheet if none specified
	kwargs = {"engine": "openpyxl"}
	if sheet_name is not None:
		kwargs["sheet_name"] = sheet_name
	return pd.read_excel(excel_io, **kwargs)


@st.cache_data(show_spinner=False)
def get_sheet_names(file_bytes: bytes) -> List[str]:
	from openpyxl import load_workbook
	excel_io = BytesIO(file_bytes)
	wb = load_workbook(excel_io, read_only=True, data_only=True)
	sheets = wb.sheetnames
	wb.close()
	return sheets


def detect_column_types(df: pd.DataFrame) -> Tuple[List[str], List[str], List[str]]:
	"""Return (datetime_cols, numeric_cols, categorical_cols)."""
	# Try to coerce likely date columns
	for col in df.columns:
		if df[col].dtype == object and any(k in str(col).lower() for k in ["date", "jour", "journee", "timestamp", "time"]):
			with pd.option_context('mode.chained_assignment', None):
				try:
					df[col] = pd.to_datetime(df[col], errors="coerce")
				except Exception:
					pass

	datetime_cols = [c for c in df.columns if pd.api.types.is_datetime64_any_dtype(df[c])]
	numeric_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
	# Categorical: heuristic
	categorical_cols = [
		c for c in df.columns
		if c not in datetime_cols and (df[c].dtype == object or df[c].nunique(dropna=True) <= max(50, int(0.05 * len(df))))
	]
	return datetime_cols, numeric_cols, categorical_cols


def apply_filters(df: pd.DataFrame, date_col: str | None, categorical_cols: List[str]) -> pd.DataFrame:
	filtered = df.copy()
	if date_col:
		min_dt = pd.to_datetime(filtered[date_col]).min()
		max_dt = pd.to_datetime(filtered[date_col]).max()
		if pd.isna(min_dt) or pd.isna(max_dt):
			pass
		else:
			start, end = st.sidebar.date_input(
				"Plage de dates",
				value=(min_dt.date(), max_dt.date()),
				min_value=min_dt.date(),
				max_value=max_dt.date(),
			)
			if isinstance(start, tuple):
				start, end = start  # streamlit older versions
			mask = (pd.to_datetime(filtered[date_col]) >= pd.to_datetime(start)) & (pd.to_datetime(filtered[date_col]) <= pd.to_datetime(end))
			filtered = filtered.loc[mask]

	for col in categorical_cols[:8]:  # limit to avoid too many widgets
		values = filtered[col].dropna().astype(str).unique().tolist()
		values.sort()
		selected = st.sidebar.multiselect(f"Filtrer {col}", values, default=values)
		if selected and len(selected) < len(values):
			filtered = filtered[filtered[col].astype(str).isin(selected)]

	return filtered


def kpi_section(df: pd.DataFrame, numeric_cols: List[str]) -> None:
	st.subheader("Indicateurs clés (KPI)")
	if not numeric_cols:
		st.info("Aucune colonne numérique détectée pour les KPI.")
		return
	metric_col = st.selectbox("Colonne de mesure", numeric_cols, index=0, key="kpi_metric")
	total = float(df[metric_col].sum())
	avg = float(df[metric_col].mean()) if len(df) else 0.0
	count_rows = int(len(df))
	left, mid, right = st.columns(3)
	left.metric("Somme", f"{total:,.2f}")
	mid.metric("Moyenne", f"{avg:,.2f}")
	right.metric("Nombre de lignes", f"{count_rows:,}")


def time_series_section(df: pd.DataFrame, datetime_cols: List[str], numeric_cols: List[str]) -> None:
	st.subheader("Tendance temporelle")
	if not datetime_cols or not numeric_cols:
		st.info("Sélectionnez au moins une colonne date et une colonne numérique.")
		return
	date_col = st.selectbox("Colonne de date", datetime_cols, index=0)
	value_col = st.selectbox("Mesure", numeric_cols, index=0, key="ts_value")
	agg_label = st.selectbox("Agrégateur", ["Somme", "Moyenne", "Compte"], index=0)
	freq_label = st.selectbox("Fréquence", ["Quotidien", "Hebdomadaire", "Mensuel", "Trimestriel"], index=2)
	freq_map = {"Quotidien": "D", "Hebdomadaire": "W-MON", "Mensuel": "MS", "Trimestriel": "QS"}
	agg_map = {"Somme": "sum", "Moyenne": "mean", "Compte": "count"}
	df2 = df.dropna(subset=[date_col]).copy()
	df2[date_col] = pd.to_datetime(df2[date_col])
	grouped = df2.set_index(date_col).resample(freq_map[freq_label])[value_col].agg(agg_map[agg_label]).reset_index()
	grouped.rename(columns={value_col: agg_label}, inplace=True)
	fig = px.line(grouped, x=date_col, y=agg_label, markers=True, title=f"{agg_label} de {value_col} ({freq_label})")
	st.plotly_chart(fig, use_container_width=True)


def category_breakdown_section(df: pd.DataFrame, categorical_cols: List[str], numeric_cols: List[str]) -> None:
	st.subheader("Répartition par catégorie")
	if not categorical_cols or not numeric_cols:
		st.info("Sélectionnez une colonne catégorielle et une colonne numérique.")
		return
	cat_col = st.selectbox("Dimension", categorical_cols, index=0)
	value_col = st.selectbox("Mesure", numeric_cols, index=min(1, len(numeric_cols)-1), key="cat_value")
	agg_label = st.selectbox("Agrégateur", ["Somme", "Moyenne", "Compte"], index=0, key="cat_agg")
	agg_map = {"Somme": "sum", "Moyenne": "mean", "Compte": "count"}
	grouped = df.groupby(cat_col, dropna=False)[value_col].agg(agg_map[agg_label]).reset_index()
	grouped.rename(columns={value_col: agg_label}, inplace=True)
	grouped = grouped.sort_values(agg_label, ascending=False).head(25)
	c1, c2 = st.columns(2)
	fig_bar = px.bar(grouped, x=cat_col, y=agg_label, title=f"{agg_label} de {value_col} par {cat_col}")
	c1.plotly_chart(fig_bar, use_container_width=True)
	fig_pie = px.pie(grouped, names=cat_col, values=agg_label, title=f"Part de {value_col} par {cat_col}", hole=0.4)
	c2.plotly_chart(fig_pie, use_container_width=True)


def histogram_section(df: pd.DataFrame, numeric_cols: List[str]) -> None:
	st.subheader("Distribution (Histogramme)")
	if not numeric_cols:
		st.info("Aucune colonne numérique disponible.")
		return
	value_col = st.selectbox("Colonne numérique", numeric_cols, index=0, key="hist_value")
	bins = st.slider("Nombre de classes (bins)", min_value=5, max_value=100, value=30)
	fig = px.histogram(df, x=value_col, nbins=bins, title=f"Histogramme de {value_col}")
	st.plotly_chart(fig, use_container_width=True)


def data_preview_and_download(df: pd.DataFrame):
	st.subheader("Aperçu des données filtrées")
	st.dataframe(df.head(1000))
	csv = df.to_csv(index=False).encode("utf-8")
	st.download_button("Télécharger CSV filtré", data=csv, file_name="donnees_filtrees.csv", mime="text/csv")


# --------------- UI ---------------

def main() -> None:
	st.title("Tableau de bord interactif à partir d'Excel")
	st.caption("Importez votre fichier Excel et construisez des graphiques en fonction des statistiques qu'il contient.")

	ensure_deps()

	with st.sidebar:
		st.header("Source des données")
		uploaded = st.file_uploader("Fichier Excel (.xlsx)", type=["xlsx"]) 
		path_text = st.text_input("ou chemin absolu du fichier sur le serveur", value="")
		read_btn = st.button("Charger le fichier")

	file_bytes: bytes | None = None
	if uploaded is not None:
		file_bytes = uploaded.getvalue()
	elif path_text and os.path.exists(path_text) and read_btn:
		with open(path_text, "rb") as f:
			file_bytes = f.read()

	if file_bytes is None:
		st.info("Chargez un fichier Excel via la barre latérale pour commencer.")
		return

	sheets = get_sheet_names(file_bytes)
	sheet = st.sidebar.selectbox("Feuille", sheets, index=0)

	df = read_excel_to_dataframe(file_bytes, sheet)
	# Clean column names a little
	df.columns = [str(c).strip().replace("\n", " ") for c in df.columns]

	dt_cols, num_cols, cat_cols = detect_column_types(df)

	st.sidebar.header("Filtres")
	date_col = st.sidebar.selectbox("Colonne de date", ["(aucune)"] + dt_cols, index=0)
	date_col = None if date_col == "(aucune)" else date_col

	df_filtered = apply_filters(df, date_col, cat_cols)

	# Sections
	kpi_section(df_filtered, num_cols)
	st.divider()
	time_series_section(df_filtered, [c for c in dt_cols if c == date_col] if date_col else dt_cols, num_cols)
	st.divider()
	category_breakdown_section(df_filtered, cat_cols, num_cols)
	st.divider()
	histogram_section(df_filtered, num_cols)
	st.divider()
	data_preview_and_download(df_filtered)


if __name__ == "__main__":
	main()
